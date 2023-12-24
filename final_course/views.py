import datetime
import json
import logging
import re

import openpyxl
from django.utils import timezone
from django.views.decorators.cache import cache_page
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from openpyxl.styles import Font

from final_course import utils


from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User
from django.http import HttpRequest, HttpResponse, Http404
from django.shortcuts import render
from django.core.validators import EmailValidator

from rest_framework import status
from rest_framework.decorators import (
    api_view,
    permission_classes,
)
from rest_framework.generics import get_object_or_404
from rest_framework.permissions import IsAuthenticated, AllowAny, IsAdminUser
from rest_framework.request import Request
from rest_framework.response import Response

from final_course import serializers, models, permission


logger = logging.getLogger(__name__)

MESSAGE = {
    "message_created": "Успешно создана!",
    "message_updated": "Успешно обновлен!",
    "message_deleted": "Успешно удален!",
    "message_error": "Ошибка",
}


def home(request: HttpRequest) -> HttpResponse:
    """
    Этот контроллер отвечает за вход на главную страницу!

    :param request: HttpRequest: Get the request object
    :return: A HttpResponse object
    :doc-author: Dias
    """

    return render(request=request, template_name="home.html")


@swagger_auto_schema(
    method="POST",
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            "username": openapi.Schema(
                type=openapi.TYPE_STRING, description="Имя пользователя на английском!"
            ),
            "password": openapi.Schema(
                type=openapi.TYPE_STRING,
                description="Пароль по требованию: мин. 8 символов, должен содержать хотя бы одну большую букву и "
                "маленькую, должен содержать цифру и спец. символы #?!@$%^&*-",
            ),
            "email": openapi.Schema(
                type=openapi.TYPE_STRING, description="Почта пример: ex@yourmail.com"
            ),
        },
        required=["username", "password", "email"],
    ),
    responses={
        201: MESSAGE["message_created"],
        400: "Не все поля заполнены / Пользователь с таким email уже существует! / Не правильный email или пароль",
    },
)
@api_view(http_method_names=["POST"])
@permission_classes([AllowAny])
def register(request: Request) -> Response:
    """
    Контроллер register используется для регистрации нового пользователя.

    Args:
        request (Request): Объект запроса, ожидается метод POST с данными пользователя в request.body.

    Returns:
        Response: Объект ответа с результатом регистрации и соответствующим статус-кодом.

    Raises:
        None.
    """
    if request.method == "POST":
        form = json.loads(request.body)
        username = form.get("username")
        password = form.get("password")
        email = form.get("email")
        regex = r"^(?=.*?[a-z])(?=.*?[A-Z])(?=.*?[0-9])(?=.*?[#?!@$%^&*-]).{8,16}$"

        if not (username and password and email):
            return Response(
                data={"error": "Не все поля заполнены."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if User.objects.filter(email=email).exists():
            return Response(
                data={"error": "Пользователь с таким email уже существует."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            if EmailValidator(email) and re.match(regex, password):
                User.objects.create(
                    username=username,
                    password=make_password(password),
                    email=email,
                )
                return Response(
                    data=MESSAGE["message_created"],
                    status=status.HTTP_201_CREATED,
                )
        except ValueError as e:
            return Response(
                data={"error": f"Неправильный email {email} или пароль {password}"},
                status=status.HTTP_400_BAD_REQUEST,
            )


@swagger_auto_schema(
    method="GET",
    responses={
        200: serializers.UserSerializer,
        401: utils.res_401,
        405: "Method not allowed",
        404: "Пользователь не найден! (User not found)",
    },
    security=[{"Bearer": []}],
)
@api_view(http_method_names=["GET"])
@permission_classes([IsAuthenticated])
def login_user(request: Request, pk) -> Response:
    """
    Получение данных пользователя по его идентификатору. (Authentificated)

        Parameters:
            request (Request): Объект запроса.
            pk (int): Идентификатор пользователя.

        Returns:
            Response: Объект ответа с данными пользователя или сообщением об ошибке.

        Raises:
            Http404: Если пользователь с указанным идентификатором не найден.
    """
    try:
        if request.method == "GET":
            user = User.objects.get(id=int(pk))
            data = serializers.UserSerializer(user).data
            return Response(data=data, status=status.HTTP_200_OK)
        else:
            return Response(
                data={"message": "Method not allowed"},
                status=status.HTTP_405_METHOD_NOT_ALLOWED,
            )
    except Http404:
        return Response(
            data={"message": "User not found"}, status=status.HTTP_404_NOT_FOUND
        )


@swagger_auto_schema(
    methods=["POST", "PUT", "DELETE"],
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            "first_name": openapi.Schema(
                type=openapi.TYPE_STRING, description="Имя кандидата на английском!"
            ),
            "last_name": openapi.Schema(
                type=openapi.TYPE_STRING, description="Фамилия кандидата на английском!"
            ),
            "middle_name": openapi.Schema(
                type=openapi.TYPE_STRING,
                description="Отчество кандидата на английском!",
            ),
            "was_born": openapi.Schema(
                type=openapi.TYPE_STRING,
                description="Дата рождения кандидата (формат DD-MM-YYYY)!",
            ),
            "phone_number": openapi.Schema(
                type=openapi.TYPE_STRING, description="Номер телефона кандидата!"
            ),
            "photo": openapi.Schema(
                type=openapi.TYPE_STRING, description="Ссылка на фото кандидата!"
            ),
            "where_from": openapi.Schema(
                type=openapi.TYPE_STRING,
                description="Место проживания кандидата (откуда (город))!",
            ),
            "where_to": openapi.Schema(
                type=openapi.TYPE_STRING,
                description="Место проживания кандидата (адрес)!",
            ),
            "job_title": openapi.Schema(
                type=openapi.TYPE_STRING,
                description="Название вакансии кандидата! (ВАКАНСИЯ ДОЛЖЕН БЫТЬ В БД)",
            ),
            "job_xp": openapi.Schema(
                type=openapi.TYPE_STRING, description="Опыт работы кандидата!"
            ),
            "education": openapi.Schema(
                type=openapi.TYPE_STRING, description="Уровень образования кандидата!"
            ),
            "skills": openapi.Schema(
                type=openapi.TYPE_STRING, description="Навыки кандидата!"
            ),
            "languages": openapi.Schema(
                type=openapi.TYPE_STRING, description="Знание языков кандидата!"
            ),
            "about_me": openapi.Schema(
                type=openapi.TYPE_STRING, description="Краткое описание кандидата!"
            ),
        },
        required=[
            "first_name",
            "last_name",
            "phone_number",
            "where_from",
            "job_title",
            "was_born",
        ],
    ),
    responses={
        200: MESSAGE["message_deleted"],
        403: "Резюме с такими данными уже существует.",
        201: MESSAGE["message_created"],
        400: "Аргумент не поддерживается!",
        500: "Internal Server Error",
    },
    security=[{"Bearer": []}],
)
@api_view(http_method_names=["POST", "PUT", "DELETE"])
@permission_classes([AllowAny])
def cv(request: Request, pk=None) -> Response:
    """
    Контроллер cv используется для создания, обновления и удаления резюме.

    Parameters:
        request (Request): Объект запроса.
        pk (int, optional): Идентификатор резюме (по умолчанию None).

    Returns:
        Response: Объект ответа с результатом операции и соответствующим статус-кодом.

    Raises:
        None.
    """
    if request.method == "POST":
        # print(request.data)
        # print(request.data.get("photo"))
        # print(request.FILES.get("photo"))
        form = request.data
        first_name = form.get("name")
        last_name = form.get("lastName")
        middle_name = form.get("middleName")
        was_born = form.get("wasBorn")
        phone_number = form.get("phone_number")
        photo = form.get("photo")
        where_from = form.get("whereFrom")
        where_to = form.get("whereTo")
        job_title = models.Vacancy.objects.get(title=form.get("jobTitle"))
        job_xp = form.get("jobXp")
        education = form.get("education")
        skills = form.get("skills")
        languages = form.get("languages")
        about_me = form.get("aboutMe")

        if models.CV.objects.filter(
            last_name=last_name, first_name=first_name, middle_name=middle_name
        ).exists():
            return Response(
                data={"error": "Резюме с такими данными уже существует."},
                status=status.HTTP_403_FORBIDDEN,
            )

        models.CV.objects.create(
            first_name=first_name,
            last_name=last_name,
            middle_name=middle_name,
            was_born=was_born,
            phone_number=phone_number,
            photo=photo,
            where_from=where_from,
            where_to=where_to,
            job_title=job_title,
            job_xp=job_xp,
            education=education,
            skills=skills,
            languages=languages,
            about_me=about_me,
        )

        return Response(
            data=MESSAGE["message_created"],
            status=status.HTTP_201_CREATED,
        )
    elif request.method == "PUT":
        if pk:
            get_cv = get_object_or_404(models.CV, id=int(pk))
            get_cv.is_viewed = True
            if request.data.get("is_accept"):
                get_cv.is_accept = True
            if request.data.get("is_declined"):
                get_cv.is_declined = True
            try:
                get_cv.save()
                return Response(
                    data=MESSAGE["message_updated"], status=status.HTTP_200_OK
                )
            except Exception as e:
                return Response(
                    data={"message": str(e)}, status=status.HTTP_400_BAD_REQUEST
                )
        else:
            return Response(
                data={"Аргумент не поддерживается!"}, status=status.HTTP_400_BAD_REQUEST
            )
    elif request.method == "DELETE":
        try:
            get_object_or_404(models.CV, id=int(pk)).delete()
            return Response(data=MESSAGE["message_deleted"], status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                data={"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


@swagger_auto_schema(
    method="GET",
    responses={200: serializers.VacanciesSerializer, 500: "INTERNAL SERVER ERROR!"},
    security=[{"Bearer": []}],
)
@api_view(http_method_names=["GET"])
@permission_classes([AllowAny])
def vacancy(request: Request) -> Response:
    """
    Получение списка всех вакансий.

    Parameters:
        request (Request): Объект запроса.

    Returns:
        Response: Объект ответа со списком активных вакансий или сообщением об ошибке.
    """
    if request.method == "GET":
        # print(request.user)
        try:
            if (
                request.user.is_superuser or request.user.is_staff
            ):  # Надо сделать отдельный маршрут с правами(permissions)
                vacations = models.Vacancy.objects.all()
                print("SU or Staff")
            else:
                vacations = models.Vacancy.objects.filter(is_active=True)
                print("anonym")
            serializer = serializers.VacanciesSerializer(vacations, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                data={"error": f"Произошла ошибка: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


@swagger_auto_schema(
    method="GET",
    responses={200: serializers.CVSerializer, 500: "INTERNAL SERVER ERROR!"},
    security=[{"Bearer": []}],
)
@api_view(http_method_names=["GET"])
@permission_classes([IsAuthenticated, permission.IsInGroupPermission])
# @cache_page(30 * 60)
def all_cv_boss(request: Request) -> Response:
    """
    Показывает все резюме только пользователям с правами в группе "Директор / Бухгалтерия".

    Args:
        request (Request): Объект запроса.

    Returns:
        Response: Объект ответа со списком резюме или сообщением об ошибке.
    """
    try:
        serializer = serializers.CVSerializer(models.CV.objects.all(), many=True)
        # print(serializer.data[0])
        return Response(data=serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        return Response(
            data={"message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@swagger_auto_schema(
    method="GET",
    responses={200: serializers.CVSerializer, 500: "INTERNAL SERVER ERROR!"},
    security=[{"Bearer": []}],
)
@api_view(http_method_names=["GET"])
@permission_classes([IsAuthenticated, permission.IsMaster])
# @cache_page(30 * 60)
def all_cv(request: Request) -> Response:
    """
    Показывает все резюме только у кого есть права (Группы - Начальники объекта, Мастера)!

    Args:
        request (Request): Объект запроса.

    Returns:
        Response: Объект ответа со списком резюме или сообщением об ошибке.
    """
    try:
        serializer = serializers.CVSerializer(
            models.CV.objects.all().order_by("rating"), many=True
        )
        # print(serializer.data[0]
        return Response(data=serializer.data, status=status.HTTP_200_OK)
    except Exception as e:
        return Response(
            data={"message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@swagger_auto_schema(
    methods=["POST", "PUT", "DELETE", "PATCH"],
    request_body=serializers.VacanciesSerializer,
    responses={
        200: MESSAGE["message_deleted"],
        403: "Резюме с такими данными уже существует.",
        201: MESSAGE["message_created"] + " / " + MESSAGE["message_updated"],
        400: "All fields are required",
        500: "Internal Server Error",
    },
    security=[{"Bearer": []}],
)
@api_view(http_method_names=["POST", "PUT", "PATCH", "DELETE"])
@permission_classes([IsAuthenticated, IsAdminUser])
def create_vacancy(request: Request, pk=None) -> Response:
    """
    Создание, обновление, удаление вакансии! Доступно только SU & администраторам.

    Args:
        request (Request): Объект запроса.
        pk (int, optional): Идентификатор вакансии (по умолчанию None).

    Returns:
        Response: Объект ответа с результатом операции и соответствующим статус-кодом.

    """
    form = request.data
    print(form)
    title = form.get("title")
    description = form.get("description")
    is_active = form.get("is_active")
    required_person = form.get("required_person")

    if request.method == "POST":
        if not title or not description or not required_person:
            return Response(
                data={"message": "All fields are required"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            models.Vacancy.objects.create(
                title=title, description=description, required_person=required_person
            )
            return Response(
                data=MESSAGE["message_created"], status=status.HTTP_201_CREATED
            )
        except Exception as e:
            return Response(
                data={"error": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )
    if pk:
        get_vacancy = models.Vacancy.objects.get(id=int(pk))
        if request.method == "PUT" or request.method == "PATCH":
            if not title or not description or not required_person:
                return Response(
                    data={"message": "All fields are required"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            get_vacancy.title = title
            get_vacancy.description = description
            get_vacancy.is_active = is_active
            get_vacancy.required_person = required_person

            try:
                get_vacancy.save()
                return Response(
                    data=MESSAGE["message_updated"], status=status.HTTP_201_CREATED
                )
            except Exception as e:
                return Response(
                    data={"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        elif request.method == "DELETE":
            try:
                get_vacancy.delete()
                return Response(
                    data=MESSAGE["message_deleted"], status=status.HTTP_200_OK
                )
            except Exception as e:
                return Response(
                    data={"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )


@api_view(http_method_names=["GET"])
@permission_classes([IsAuthenticated, IsAdminUser])
def like_dislike_cv_get(request: Request, pk: str) -> Response:
    if request.method == "GET":
        data = {
            "like": models.Like.objects.filter(which_cv__id=int(pk))
            .filter(status=True)
            .count(),
            "dislike": models.Like.objects.filter(which_cv__id=int(pk))
            .filter(status=False)
            .count(),
        }
        print(data)
        return Response(data=data, status=status.HTTP_200_OK)


@api_view(http_method_names=["GET", "POST"])
@permission_classes([IsAuthenticated, permission.IsMaster])
def like_dislike_cv_post(request: Request, pk: str, st: str = None) -> Response:
    user_ = request.user
    cv_obj = models.CV.objects.get(id=int(pk))
    if request.method == "GET":
        if models.Like.objects.get(which_cv=cv_obj, clicked_by=user_):
            return Response(
                data={"message": "Объект оценен пользователем!"},
                status=status.HTTP_202_ACCEPTED,
            )
        else:
            return Response(
                data={"message": "Объект не оценен или не существует!"},
                status=status.HTTP_204_NO_CONTENT,
            )
    elif request.method == "POST":
        try:
            cv_rating = models.Like.objects.get(which_cv=cv_obj, clicked_by=user_)
        except models.Like.DoesNotExist:
            models.Like.objects.create(
                which_cv=cv_obj, clicked_by=user_, status=bool(st)
            )
        else:
            if (bool(st) is True and cv_rating.status is True) or (
                bool(st) is False and cv_rating.status is False
            ):
                cv_rating.delete()
            elif (bool(st) is True and cv_rating.status is not True) or (
                bool(st) is False and cv_rating.status is not False
            ):
                cv_rating.delete()
            else:
                cv_rating.status = bool(st)
                cv_rating.save()
        return Response(data=MESSAGE["message_created"], status=status.HTTP_201_CREATED)


@api_view(http_method_names=["GET"])
@permission_classes([AllowAny])
def news_to_front(request: Request) -> Response:
    """
    Получение новости из стороннего сайта!

    :param request: Получает объект запроса
    :return: Кэшированную новость
    """
    get_cached_news = utils.CustomCache.get_cache(key="news")
    if get_cached_news:
        return Response(data=get_cached_news, status=status.HTTP_200_OK)
    else:
        try:
            news = utils.news()
            utils.CustomCache.set_cache(key="news", data=news, timeout=3600)
            return Response(data=news, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                data={"message": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


def generate_xlsx_data(
    data, date_filter: datetime.date | list[datetime.date, datetime.date]
):
    """
    Подгатавливает файл .xlsx записав данные!

    :param data: Принимает сериализованный объект БД
    :param date_filter: Принимает дату или список с датами(2)
    :return: Возвращает записанный файл или "Нет резюме за этот период" в файле! То есть в любом случае возвращает файл!
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(
        [
            "Имя",
            "Фамилия",
            "Номер телефона",
            "Откуда",
            "На должность",
            "Когда отправлено",
        ]
    )

    if isinstance(date_filter, datetime.date):
        filtered_cv = [
            cv
            for cv in data
            if date_filter
            <= datetime.datetime.strptime(cv["created_at"], "%Y-%m-%d").date()
            <= timezone.now().date()
        ]
        if filtered_cv:
            print(filtered_cv)
            for cv in filtered_cv:
                ws.append(list(cv.values()))

            row = ws.row_dimensions[1]
            row.font = Font(size=14, bold=True, color="00FFFF00")
            return wb
        else:
            print("else")
            ws["A2"] = "Нет резюме за этот период"
            ws.merge_cells("A2:F2")
            return wb
    if isinstance(date_filter, list):
        filtered_cv = [
            cv
            for cv in data
            if date_filter[0]
            <= datetime.datetime.strptime(cv["created_at"], "%Y-%m-%d")
            <= date_filter[1]
        ]
        if filtered_cv:
            for cv in filtered_cv:
                ws.append(list(cv.values()))

            row = ws.row_dimensions[1]
            row.font = Font(size=14, bold=True, color="00FFFF00")
            return wb
        else:
            ws["A2"] = "Нет резюме за этот период"
            ws.merged_cells("A2:F2")
            return wb
    if date_filter is None:
        for cv in data:
            ws.append(list(cv.values()))
        row = ws.row_dimensions[1]
        row.font = Font(size=14, bold=True, color="00FFFF00")
        return wb


@permission_classes([IsAuthenticated, permission.IsInGroupPermission])
def download_xlsx_period(request, period=None):
    """
    Подгатавливает данные к скачиванию
    :param request: Запрос
    :param period: Строка превращается в объект date (today, week, month)
    :return: Файл для скачивания
    """
    allcv = utils.CustomCache.get_cache(key="allCv")

    if allcv is None:
        allcv = serializers.CVSerializerDownload(
            models.CV.objects.all(), many=True
        ).data
        utils.CustomCache.set_cache(key="allCv", data=allcv, timeout=1800)

    res = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # Функция для получения правильного названия файла в зависимости от периода
    def get_filename(period_str):
        return f'attachment; filename="Резюме за {period_str}.xlsx"'

    if period == "today":
        date_filter = timezone.now().date()
        # print(date_filter, "today", type(date_filter))
    elif period == "month":
        date_filter = timezone.now().date() - datetime.timedelta(days=30)
    elif period == "week":
        date_filter = timezone.now().date() - datetime.timedelta(weeks=1)
    else:
        date_filter = None

    wb = generate_xlsx_data(data=allcv, date_filter=date_filter)

    if wb:
        wb.save(res)
        res["Content-Disposition"] = get_filename(period)
        return res
    else:
        return HttpResponse(
            content=json.dumps({"status": f"Нет резюме за {period}"}),
            status=204,
            content_type="application/json",
        )


@permission_classes([IsAuthenticated, permission.IsInGroupPermission])
def download_xlsx_custom(request):
    """
    Подгатавливает данные к скачиванию!
    Из request получает startDate & endDate! Период времени!

    :param request: Запрос
    :return: Файл для скачивания
    """
    start = datetime.datetime.strptime(request.GET.get("startDate"), "%Y-%m-%d")
    end = datetime.datetime.strptime(request.GET.get("endDate"), "%Y-%m-%d")
    allcv = utils.CustomCache.get_cache(key="allCv")

    res = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    if allcv is None:
        allcv = serializers.CVSerializerDownload(
            models.CV.objects.all(), many=True
        ).data
        utils.CustomCache.set_cache(key="allCv", data=allcv, timeout=1800)

    wb = generate_xlsx_data(allcv, date_filter=[start, end])

    if wb:
        wb.save(res)
        res[
            "Content-Disposition"
        ] = f'attachment; filename="Резюме за {start} по {end}.xlsx"'
        return res
    else:
        return HttpResponse(
            content=json.dumps({"status": f"Нет резюме за выбранный период"}),
            status=204,
            content_type="application/json",
        )


@api_view(http_method_names=["GET", "PATCH"])
@permission_classes([IsAuthenticated, permission.IsInGroupPermission])
def applications(request: Request) -> Response:
    data = utils.CustomCache.get_cache(key="applications")
    if data is None:
        data = utils.CustomCache.set_cache(
            key="applications",
            data=serializers.ApplicationSerializer(
                models.Application.objects.all(), many=True
            ).data,
            timeout=1000,
        )
    if request.method == "GET":
        return Response(data=data, status=status.HTTP_200_OK)
    elif request.method == "PATCH":
        status_to_db = request.GET.get("status")
        pk = request.GET.get("pk")
        try:
            get_bid = models.Application.objects.get(id=int(pk))
            get_bid.is_accepted = status_to_db
            get_bid.save()
            return Response(
                data=MESSAGE["message_updated"], status=status.HTTP_201_CREATED
            )
        except Exception as e:
            return Response(
                data={"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


@api_view(http_method_names=["GET", "POST"])
@permission_classes([IsAuthenticated, permission.IsMaster])
def application(request: Request) -> Response:
    data = utils.CustomCache.get_cache(key="application")
    if data is None:
        data = utils.CustomCache.set_cache(
            key="application",
            data=serializers.ApplicationSerializer(
                models.Application.objects.filter(_user=request.user), many=True
            ).data,
            timeout=1000,
        )
    if request.method == "GET":
        return Response(data=data, status=status.HTTP_200_OK)
    elif request.method == "POST":
        form = request.data
        # print(form)
        diameter = int(form.get("diameter"))
        thickness = int(form.get("thickness"))
        length = int(form.get("length"))
        type_of = form.get("type_of")
        object_to = form.get("object_to")

        if (diameter or thickness or length or type_of or object_to) is None:
            return Response(
                data={"message": "Заполните все поля"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            models.Application.objects.create(
                diameter=diameter,
                thickness=thickness,
                type_of=type_of,
                length=length,
                _user=request.user,
                object_to=object_to,
            )
            return Response(
                data=MESSAGE["message_created"], status=status.HTTP_201_CREATED
            )
        except ValueError:
            return Response(
                data={"message": "Отправлены неправильные поля"},
                status=status.HTTP_400_BAD_REQUEST,
            )


@api_view(http_method_names=["GET", "PATCH"])
@permission_classes([IsAuthenticated])
def profile(request: Request) -> Response:
    user_id = int(request.GET.get("userId"))
    # print(user_id)

    user = User.objects.get(id=user_id)
    # print(user)

    prof = models.UserProfile.objects.get(user=user)
    # print(prof)

    data = serializers.UserProfileSerializer(prof, many=False).data
    # print(data)
    if request.method == "GET":
        try:
            return Response(data=data, status=status.HTTP_200_OK)
        except Exception as e:
            return Response(
                data={"error": str(e)},
                status=status.HTTP_404_NOT_FOUND,
            )

    elif request.method == "PATCH":
        form = request.data
        # print(form)

        first_name = form.get("first_name")
        last_name = form.get("last_name")
        was_born = form.get("was_born")
        phone_number = form.get("phone_number")
        avatar = form.get("avatar")

        if (first_name and last_name and was_born and phone_number and avatar) is None:
            return Response(
                data={"message": "Заполните хотя бы одно поле"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        else:
            user.first_name = first_name
            user.last_name = last_name
            prof.was_born = was_born
            prof.phone_number = phone_number
            prof.avatar = avatar
            try:
                user.save()
                prof.save()
                return Response(
                    data=MESSAGE["message_updated"], status=status.HTTP_201_CREATED
                )
            except Exception as e:
                return Response(
                    data=str(e), status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
